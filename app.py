from flask import Flask, request, render_template, redirect, url_for, session, jsonify, send_file
from flask import Flask, render_template, request, redirect, url_for, send_file, session
from models import db, Dispatch, Center, SmsTemplate, Notice
import pandas as pd
import io
import os  # 💡 파일 경로 생성을 위해 추가
from datetime import datetime, timedelta
import urllib.parse
import requests
import urllib3
import json
import concurrent.futures
import math  # 💡 [신규] 최적 경로 거리 계산을 위한 수학 모듈
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from werkzeug.utils import secure_filename  # 💡 파일명 안전 처리를 위해 추가
from models import db, Dispatch, Center, SmsTemplate, Notice, SystemSettings
from io import BytesIO

# 💡 [신규] 두 좌표 간의 거리를 계산하는 하버사인 알고리즘
def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    dlat = math.radians(float(lat2) - float(lat1))
    dlon = math.radians(float(lon2) - float(lon1))
    a = math.sin(dlat / 2)**2 + math.cos(math.radians(float(lat1))) * math.cos(math.radians(float(lat2))) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c
    
# 💡 [신규] 카카오 주소 -> 좌표 변환 함수 (기존에 쓰시던 함수가 있다면 그것을 쓰셔도 됩니다)
def get_kakao_coords(address):
    KAKAO_API_KEY = "f70047282a8b7f30cd02fd2cfc00f029" 
    url = "https://dapi.kakao.com/v2/local/search/address.json"
    headers = {"Authorization": f"KakaoAK {KAKAO_API_KEY}"}
    params = {"query": address}
    try:
        res = requests.get(url, headers=headers, params=params)
        if res.status_code == 200:
            documents = res.json().get('documents')
            if documents:
                return documents[0]['x'], documents[0]['y'] # (경도, 위도) 반환
    except Exception as e:
        print(f"카카오 API 변환 에러: {e}")
    return None, None

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'jette_super_secret_admin_key' 

# 💡 [신규] 이미지 업로드용 서버 설정 추가
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 최대 16MB 업로드 제한

# 업로드 폴더 자동 생성
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# 이미지 확장자 검증 함수
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db.init_app(app)
with app.app_context():
    db.create_all()

# (이하 기존 카카오 API 및 driving_time 등 공통 함수 생략 - 파일 덮어쓰기 하셔도 모든 코드 완벽 유지됩니다.)
KAKAO_API_KEY = 'f70047282a8b7f30cd02fd2cfc00f029'
kakao_session = requests.Session()
route_time_cache = {}

def get_coords(address):
    if not address or address.strip() == '': return None, None
    encoded_address = urllib.parse.quote(address)
    url = f"https://dapi.kakao.com/v2/local/search/address.json?query={encoded_address}"
    headers = {"Authorization": f"KakaoAK {KAKAO_API_KEY}"}
    try:
        resp = kakao_session.get(url, headers=headers, verify=False, timeout=5).json()
        if resp.get('documents'): return resp['documents'][0]['x'], resp['documents'][0]['y']
    except Exception: pass
    return None, None

def get_driving_time(start_x, start_y, end_x, end_y):
    try:
        sx, sy, ex, ey = round(float(start_x), 4), round(float(start_y), 4), round(float(end_x), 4), round(float(end_y), 4)
        cache_key = f"{sx},{sy}_{ex},{ey}"
        if cache_key in route_time_cache: return route_time_cache[cache_key]
    except Exception: cache_key = None
    url = f"https://apis-navi.kakaomobility.com/v1/directions?origin={start_x},{start_y}&destination={end_x},{end_y}"
    headers = {"Authorization": f"KakaoAK {KAKAO_API_KEY}"}
    try:
        resp = kakao_session.get(url, headers=headers, verify=False, timeout=2).json()
        if resp.get('routes'): 
            duration = resp['routes'][0]['summary']['duration']
            if cache_key: route_time_cache[cache_key] = duration
            return duration
    except Exception: pass
    return 25 * 60 

def update_etas_for_driver(driver_name):
    dispatches = Dispatch.query.filter_by(driver_name=driver_name).order_by(Dispatch.delivery_seq).all()
    if not dispatches: return
    base_depart_time = dispatches[0].center_depart_time
    if not base_depart_time: return 
        
    routes_to_fetch = []
    current_x, current_y = get_coords(dispatches[0].center_address if dispatches[0].center_address else dispatches[0].store_address)
    for d in dispatches:
        if d.is_departed and d.departure_time:
            current_x, current_y = d.store_x, d.store_y
        else:
            if current_x and current_y and d.store_x and d.store_y:
                routes_to_fetch.append((current_x, current_y, d.store_x, d.store_y))
            current_x, current_y = d.store_x, d.store_y
            
    duration_map = {}
    def fetch_route_time(route):
        sx, sy, ex, ey = route
        return route, get_driving_time(sx, sy, ex, ey)
        
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        results = executor.map(fetch_route_time, list(set(routes_to_fetch)))
        for route, duration in results: duration_map[route] = duration

    current_departure = base_depart_time
    current_x, current_y = get_coords(dispatches[0].center_address if dispatches[0].center_address else dispatches[0].store_address)
    for d in dispatches:
        if d.is_departed and d.departure_time:
            current_departure = d.departure_time
            current_x, current_y = d.store_x, d.store_y
        else:
            if current_x and current_y and d.store_x and d.store_y:
                duration_sec = duration_map.get((current_x, current_y, d.store_x, d.store_y), 25 * 60)
                arrival_time = current_departure + timedelta(seconds=duration_sec)
            else:
                arrival_time = current_departure + timedelta(minutes=25)
            d.estimated_arrival = arrival_time
            current_departure = arrival_time + timedelta(minutes=d.buffer_time)
            current_x, current_y = d.store_x, d.store_y

def apply_excel_styles(worksheet, df, is_sms=False):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if is_sms and col_idx == 4:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

@app.route('/')
def home(): return redirect(url_for('admin_login'))

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == 'a13579!!':
            session['is_admin'] = True
            return redirect(url_for('admin'))
        else: return "<script>alert('비밀번호가 틀렸습니다.'); history.back();</script>"
    return '''<div style="text-align:center; margin-top:150px; font-family:'Malgun Gothic', sans-serif;"><h2 style="color:#082c84;">JETTE 관리자 로그인</h2><form method="post"><input type="password" name="password" placeholder="비밀번호 입력" required style="padding:10px; font-size:16px; border: 1px solid #ccc; border-radius: 4px;"><button type="submit" style="padding:10px 20px; font-size:16px; background:#082c84; color:white; border:none; border-radius: 4px; cursor:pointer;">접속</button></form></div>'''

@app.route('/admin_logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect(url_for('admin_login'))

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    if request.method == 'POST':
        excel_text = request.form.get('excel_text')
        if not excel_text or excel_text.strip() == '': return "입력된 데이터가 없습니다.", 400
        try:
            df = pd.read_csv(io.StringIO(excel_text), sep='\t')
            df.columns = df.columns.str.replace(' ', '')
            driver_seq_counter = {}
            address_cache = {}
            if '매장주소' in df.columns:
                unique_addresses = df['매장주소'].dropna().astype(str).str.strip().unique()
                unique_addresses = [addr for addr in unique_addresses if addr]
                def fetch_coord(addr): return addr, get_coords(addr)
                with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                    results = executor.map(fetch_coord, unique_addresses)
                    for addr, (sx, sy) in results: address_cache[addr] = (sx, sy)

            for index, row in df.iterrows():
                raw_date = str(row.get('배송일자', '')).strip()
                clean_date = raw_date.replace('년', '-').replace('월', '-').replace('일', '').replace(' ', '')
                if len(clean_date.split('-')) == 2: clean_date = f"{datetime.now().year}-{clean_date}"
                try: delivery_date = pd.to_datetime(clean_date).date()
                except: delivery_date = datetime.now().date()
                
                driver_name = str(row.get('기사명', '')).strip()
                if driver_name not in driver_seq_counter: driver_seq_counter[driver_name] = 1

                if '배송순서' in df.columns and pd.notna(row['배송순서']):
                    seq_value = int(row['배송순서'])
                    driver_seq_counter[driver_name] = seq_value + 1
                else:
                    seq_value = driver_seq_counter[driver_name]
                    driver_seq_counter[driver_name] += 1

                center_name_val = str(row.get('센터명', '')).strip()
                center_obj = Center.query.filter_by(name=center_name_val).first()
                center_addr_val = center_obj.address if center_obj else ''

                store_address_val = str(row.get('매장주소', '')).strip()
                buffer_time_val = int(row['상하차시간(분)']) if '상하차시간(분)' in df.columns and pd.notna(row['상하차시간(분)']) else 10
                sx, sy = address_cache.get(store_address_val, (None, None))

                dispatch_entry = Dispatch(
                    delivery_date=delivery_date, center_name=center_name_val, center_address=center_addr_val, 
                    vehicle_num=str(row.get('차량번호', '')).strip(), driver_name=driver_name, 
                    store_code=str(row.get('매장코드', '')).strip(), store_name=str(row.get('매장명', '')).strip(),
                    store_address=store_address_val, delivery_seq=seq_value, buffer_time=buffer_time_val, 
                    store_x=sx, store_y=sy, driver_phone=str(row.get('기사전화번호', '')).strip(),
                    store_phone=str(row.get('매장전화번호', '')).strip(), template_name=str(row.get('템플릿양식', '')).strip()
                )
                db.session.add(dispatch_entry)
            db.session.commit()
            return redirect(url_for('admin'))
        except Exception as e:
            db.session.rollback()
            return f"오류 발생: {str(e)} <br><br><a href='/admin'>돌아가기</a>"
            
    all_data = Dispatch.query.order_by(Dispatch.delivery_date, Dispatch.driver_name, Dispatch.delivery_seq).all()
    centers = Center.query.all()
    return render_template('admin.html', dispatches=all_data, centers=centers)

@app.route('/admin/add_center', methods=['POST'])
def add_center():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    name = request.form.get('center_name').strip()
    address = request.form.get('center_address').strip()
    
    # 카카오 API를 통해 주소를 좌표로 변환
    c_x, c_y = get_kakao_coords(address)
        
    if name and address:
        existing = Center.query.filter_by(name=name).first()
        if existing:
            existing.address = address
            existing.center_x = c_x
            existing.center_y = c_y
        else:
            new_center = Center(name=name, address=address, center_x=c_x, center_y=c_y)
            db.session.add(new_center)
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/delete_center/<int:center_id>', methods=['POST'])
def delete_center(center_id):
    c = Center.query.get(center_id)
    if c:
        db.session.delete(c)
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/delete_by_center', methods=['POST'])
def delete_by_center():
    center_name = request.form.get('center_name')
    if center_name:
        Dispatch.query.filter_by(center_name=center_name).delete()
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/delete_all', methods=['POST'])
def delete_all():
    db.session.query(Dispatch).delete()
    db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/delete/<int:dispatch_id>', methods=['POST'])
def delete_dispatch(dispatch_id):
    d = Dispatch.query.get(dispatch_id)
    if d:
        db.session.delete(d)
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/update_address/<int:dispatch_id>', methods=['POST'])
def update_address(dispatch_id):
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        new_address = request.form.get('new_address', '').strip()
        if new_address:
            dispatch.store_address = new_address
            sx, sy = get_coords(new_address)
            dispatch.store_x = sx
            dispatch.store_y = sy
            db.session.commit()
            update_etas_for_driver(dispatch.driver_name)
            db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/update_buffer/<int:dispatch_id>', methods=['POST'])
def update_buffer(dispatch_id):
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        new_time = request.form.get('buffer_time', type=int)
        if new_time is not None:
            dispatch.buffer_time = new_time
            db.session.commit()
            update_etas_for_driver(dispatch.driver_name) 
            db.session.commit()
    return redirect(url_for('admin'))

@app.route('/admin/update_driver/<int:dispatch_id>', methods=['POST'])
def update_driver(dispatch_id):
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        new_driver = request.form.get('new_driver_name', '').strip()
        new_vehicle = request.form.get('new_vehicle_num', '').strip()
        
        if new_driver and new_vehicle:
            old_driver = dispatch.driver_name
            dispatch.driver_name = new_driver
            dispatch.vehicle_num = new_vehicle
            if new_driver != old_driver:
                max_seq_dispatch = Dispatch.query.filter_by(driver_name=new_driver).order_by(Dispatch.delivery_seq.desc()).first()
                if max_seq_dispatch: dispatch.delivery_seq = max_seq_dispatch.delivery_seq + 1
                else: dispatch.delivery_seq = 1
            db.session.commit()
            update_etas_for_driver(old_driver)
            if new_driver != old_driver: update_etas_for_driver(new_driver)
            db.session.commit()
    return redirect(url_for('admin'))

@app.route('/driver', methods=['GET'])
def driver():
    name = request.args.get('driver_name')
    dispatches = []
    route_chunks = [] 
    date_str = ""
    active_notices = []
    
    comp_msg_setting = SystemSettings.query.filter_by(key='completion_msg').first()
    completion_message = comp_msg_setting.value if comp_msg_setting else "금일 배송도 고생 많으셨습니다!\n제때에서 발송된 카카오톡 배송승인 부탁드리겠습니다."

    if name:
        dispatches = Dispatch.query.filter_by(driver_name=name).order_by(Dispatch.delivery_seq).all()
        driver_center = dispatches[0].center_name if dispatches else ""
        
        all_active_notices = Notice.query.filter_by(is_active=True).order_by(Notice.display_seq.asc(), Notice.created_at.desc()).all()
        for n in all_active_notices:
            target_str = n.target_drivers.strip() if n.target_drivers else ""
            if "||" in target_str:
                t_center, t_drivers = target_str.split("||", 1)
            else:
                t_center = "전체"
                t_drivers = target_str
                
            if t_center != "전체" and t_center != driver_center:
                continue
                
            if not t_drivers:
                active_notices.append(n) 
            elif t_drivers.lower().startswith("contain "):
                keywords = [k.strip() for k in t_drivers[8:].split(',') if k.strip()]
                if any(k in name for k in keywords):
                    active_notices.append(n)
            elif t_drivers.lower().startswith("not contain "):
                keywords = [k.strip() for k in t_drivers[12:].split(',') if k.strip()]
                if not any(k in name for k in keywords):
                    active_notices.append(n)
            else:
                target_list = [d.strip() for d in t_drivers.split(',') if d.strip()]
                if name in target_list:
                    active_notices.append(n)
                    
        if not dispatches:
            return f"<script>alert('{name} 기사님의 배차 내역이 존재하지 않습니다.'); window.location.href='/driver';</script>"
        display_date = datetime.now().date()
        if dispatches and dispatches[0].delivery_date: display_date = dispatches[0].delivery_date
        weekdays = ['월', '화', '수', '목', '금', '토', '일']
        date_str = display_date.strftime('%y%m%d') + f"({weekdays[display_date.weekday()]})"
        valid_dispatches = [d for d in dispatches if d.store_x and d.store_y and not d.is_departed]
        chunk_size = 5
        
        # 💡 관리자 DB에 등록된 센터 좌표 조회
        target_center_obj = Center.query.filter_by(name=driver_center).first()
        if target_center_obj and target_center_obj.center_y and target_center_obj.center_x:
            cy, cx = target_center_obj.center_y, target_center_obj.center_x
        else:
            cy, cx = "37.5665", "126.9780" # 등록 안되어있을 경우 기본값(서울)
        
        for i in range(0, len(valid_dispatches), chunk_size):
            chunk = valid_dispatches[i:i+chunk_size]
            dest_d = chunk[-1]
            ep_y, ep_x = float(dest_d.store_y), float(dest_d.store_x)
            
            # 💡 출발지(sp)를 DB 센터 좌표로 완벽하게 고정
            kakaomap_app_url = f"kakaomap://route?sp={cy},{cx}&ep={ep_y},{ep_x}&by=CAR"
            
            for idx, wp in enumerate(chunk[:-1]):
                vp_key = 'vp' if idx == 0 else f"vp{idx+1}"
                kakaomap_app_url += f"&{vp_key}={float(wp.store_y)},{float(wp.store_x)}"
            
            center_addr = dispatches[0].center_address if dispatches[0].center_address else dispatches[0].store_address
            origin_encoded = urllib.parse.quote(center_addr)
            dest_encoded = urllib.parse.quote(dest_d.store_address)
            google_map_url = f"https://www.google.com/maps/dir/?api=1&origin={origin_encoded}&destination={dest_encoded}"
            waypoint_addrs = [v.store_address for v in chunk[:-1]]
            if waypoint_addrs:
                wp_encoded = urllib.parse.quote("|".join(waypoint_addrs))
                google_map_url += f"&waypoints={wp_encoded}"
                
            route_chunks.append({
                'title': f"📱 코스 ({chunk[0].delivery_seq}~{chunk[-1].delivery_seq}번)",
                'url': kakaomap_app_url, 'pc_url': google_map_url 
            })
            
    return render_template('driver.html', dispatches=dispatches, driver_name=name, route_chunks=route_chunks, date_str=date_str, active_notices=active_notices, completion_message=completion_message)

@app.route('/admin/update_phones/<int:dispatch_id>', methods=['POST'])
def update_phones(dispatch_id):
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        dispatch.driver_phone = request.form.get('driver_phone', '').strip()
        dispatch.store_phone = request.form.get('store_phone', '').strip()
        db.session.commit()
    return redirect(url_for('admin'))

@app.route('/depart_center', methods=['POST'])
def depart_center():
    driver_name = request.form.get('driver_name')
    manual_time_str = request.form.get('manual_time') 
    dispatches = Dispatch.query.filter_by(driver_name=driver_name).order_by(Dispatch.delivery_seq).all()
    if not dispatches: return "데이터 없음", 404
    depart_dt = datetime.now() 
    if manual_time_str:
        today = datetime.now().date()
        time_obj = datetime.strptime(manual_time_str, '%H:%M').time()
        depart_dt = datetime.combine(today, time_obj)
    for d in dispatches: d.center_depart_time = depart_dt
    db.session.commit()
    update_etas_for_driver(driver_name)
    db.session.commit()
    return redirect(url_for('driver', driver_name=driver_name))

@app.route('/cancel_depart', methods=['POST'])
def cancel_depart():
    driver_name = request.form.get('driver_name')
    for d in Dispatch.query.filter_by(driver_name=driver_name).all():
        d.center_depart_time = None
        d.estimated_arrival = None
    db.session.commit()
    return redirect(url_for('driver', driver_name=driver_name))

@app.route('/complete/<int:dispatch_id>', methods=['POST'])
def complete_delivery(dispatch_id):
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        dispatch.is_departed = True
        dispatch.departure_time = datetime.now()
        db.session.commit()
        update_etas_for_driver(dispatch.driver_name) 
        db.session.commit()
        return redirect(url_for('driver', driver_name=dispatch.driver_name))
    return "데이터 없음", 404

@app.route('/update_seq/<int:dispatch_id>', methods=['POST'])
def update_seq(dispatch_id):
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        new_seq = request.form.get('new_seq', type=int)
        if new_seq:
            completed_count = Dispatch.query.filter_by(driver_name=dispatch.driver_name, is_departed=True).count()
            if new_seq <= completed_count: new_seq = completed_count + 1
            dispatch.delivery_seq = new_seq
            db.session.commit()
            update_etas_for_driver(dispatch.driver_name)
            db.session.commit()
        return redirect(url_for('driver', driver_name=dispatch.driver_name))
    return "데이터 없음", 404

@app.route('/update_order', methods=['POST'])
def update_order():
    order_data = request.json
    driver_name = None
    if order_data:
        first_dispatch = Dispatch.query.get(int(order_data[0]))
        if first_dispatch:
            driver_name = first_dispatch.driver_name
            completed_count = Dispatch.query.filter_by(driver_name=driver_name, is_departed=True).count()
            for index, item_id in enumerate(order_data):
                dispatch = Dispatch.query.get(int(item_id))
                if dispatch and not dispatch.is_departed:
                    dispatch.delivery_seq = completed_count + index + 1
        db.session.commit()
        if driver_name:
            update_etas_for_driver(driver_name)
            db.session.commit()
    return {"status": "success"}

@app.route('/dashboard')
def dashboard():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    unique_centers = [c.name for c in Center.query.all()]
    all_dispatches = Dispatch.query.order_by(Dispatch.driver_name, Dispatch.delivery_seq).all()
    stats = {}
    for d in all_dispatches:
        name = d.driver_name
        if name not in stats: stats[name] = {'total': 0, 'completed': 0, 'remaining': 0, 'vehicle': d.vehicle_num, 'details': []}
        stats[name]['total'] += 1
        if d.is_departed: stats[name]['completed'] += 1
        else: stats[name]['remaining'] += 1
        stats[name]['details'].append(d)
    for name, data in stats.items():
        data['progress'] = int((data['completed'] / data['total']) * 100) if data['total'] > 0 else 0
    vehicle_stats = {
        'total': len(stats),
        'completed': sum(1 for data in stats.values() if data['remaining'] == 0),
        'pending': len(stats) - sum(1 for data in stats.values() if data['remaining'] == 0)
    }
    return render_template('dashboard.html', stats=stats, vehicle_stats=vehicle_stats, unique_centers=unique_centers)

@app.route('/download_excel')
def download_excel():
    center_filter = request.args.get('center_name', '')
    query = Dispatch.query
    if center_filter: query = query.filter_by(center_name=center_filter)
    all_data = query.order_by(Dispatch.driver_name, Dispatch.delivery_seq).all()
    
    data_list = []
    for d in all_data:
        data_list.append({
            '센터명': d.center_name, '출발센터주소': d.center_address, '배송일자': d.delivery_date,
            '차량번호': d.vehicle_num, '기사명': d.driver_name, '매장코드': d.store_code,
            '매장명': d.store_name, '도착예정시간': d.estimated_arrival.strftime('%H:%M') if d.estimated_arrival else '-',
            '실제완료시간': d.departure_time.strftime('%H:%M') if d.departure_time else '미완료',
            '센터출발시간': d.center_depart_time.strftime('%H:%M') if d.center_depart_time else '미출발',
            '완료여부': '완료' if d.is_departed else '대기', '배송순서': d.delivery_seq,
            '상하차시간(분)': d.buffer_time, '매장주소': d.store_address
        })
    df = pd.DataFrame(data_list)
    today_str = datetime.now().strftime('%Y-%m-%d')
    if all_data and all_data[0].delivery_date: today_str = all_data[0].delivery_date.strftime('%Y-%m-%d')
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='배송시간')
        worksheet = writer.sheets['배송시간']
        apply_excel_styles(worksheet, df)
        for column_cells in worksheet.columns:
            max_len = 0
            for cell in column_cells:
                if cell.value:
                    val_str = str(cell.value)
                    val_len = sum(2 if ord(c) > 127 else 1.2 for c in val_str)
                    if val_len > max_len: max_len = val_len
            worksheet.column_dimensions[column_cells[0].column_letter].width = max_len + 2
    output.seek(0)
    return send_file(output, download_name=f"{center_filter}_{today_str}_배송시간.xlsx" if center_filter else f"{today_str}_배송시간.xlsx", as_attachment=True)

@app.route('/download_template')
def download_template():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    today_str = datetime.now().strftime('%Y-%m-%d')
    example_data = {
        '센터명': ['밀양센터', '오산센터'], '배송일자': [today_str, today_str],
        '차량번호': ['임시00임 0000', '임시00임 1111'], '기사명': ['홍길동', '이순신'],
        '매장코드': ['S001', 'S002'], '매장명': ['강남A', '강남B'],
        '매장주소': ['서울특별시 강남구 테헤란로 123', '서울특별시 강남구 테헤란로 456'],
        '배송순서': [1, 2], '상하차시간(분)': [10, 10], 
        '기사전화번호': ['010-1234-5678', '010-9876-5432'], '매장전화번호': ['02-111-2222', '031-333-4444 / 010-999-8888'],
        '템플릿양식': ['A 양식', 'B 양식']
    }
    df = pd.DataFrame(example_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='업로드양식')
        worksheet = writer.sheets['업로드양식']
        apply_excel_styles(worksheet, df)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            if col_name == '배송순서': cell.comment = Comment('공란이어도 됩니다.', 'Admin')
            elif col_name == '상하차시간(분)': cell.comment = Comment('공란이어도 됩니다. (기본값: 10분)', 'Admin')
            
        for column_cells in worksheet.columns:
            max_len = 0
            for cell in column_cells:
                if cell.value:
                    val_str = str(cell.value)
                    val_len = sum(2 if ord(c) > 127 else 1.2 for c in val_str)
                    if val_len > max_len: max_len = val_len
            worksheet.column_dimensions[column_cells[0].column_letter].width = max_len + 2
    output.seek(0)
    return send_file(output, download_name="JETTE_배차업로드양식.xlsx", as_attachment=True)

@app.route('/sms')
def sms_page():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    templates = SmsTemplate.query.all()
    unique_centers = [c.name for c in Center.query.all()]
    departed_dispatches = Dispatch.query.filter(Dispatch.center_depart_time != None).order_by(Dispatch.driver_name, Dispatch.delivery_seq).all()
    return render_template('sms.html', dispatches=departed_dispatches, templates=templates, unique_centers=unique_centers)

@app.route('/sms/add_template', methods=['POST'])
def add_template():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    
    template_id = request.form.get('template_id')
    name = request.form.get('template_name').strip()
    subject = request.form.get('template_subject').strip()
    sender = request.form.get('template_sender').strip()
    content = request.form.get('template_content').strip()
    
    if name and content and subject and sender:
        if template_id:
            existing = SmsTemplate.query.get(template_id)
            if existing:
                existing.name = name
                existing.subject = subject
                existing.sender_phone = sender
                existing.content = content
        else:
            existing = SmsTemplate.query.filter_by(name=name).first()
            if existing:
                existing.subject = subject
                existing.sender_phone = sender
                existing.content = content
            else:
                db.session.add(SmsTemplate(name=name, subject=subject, sender_phone=sender, content=content))
        db.session.commit()
    return redirect(url_for('sms_page'))

@app.route('/sms/delete_template/<int:template_id>', methods=['POST'])
def delete_template(template_id):
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    t = SmsTemplate.query.get(template_id)
    if t:
        db.session.delete(t)
        db.session.commit()
    return redirect(url_for('sms_page'))

@app.route('/download_sms_excel')
def download_sms_excel():
    if not session.get('is_admin'): 
        return redirect(url_for('admin_login'))
    
    center_name = request.args.get('center_name', '')
    
    # 💡 HTML(프론트엔드)에서 넘겨준 토글(ON/OFF) 값 판단
    filter_past = request.args.get('filter_past', 'true') == 'true'
    now = datetime.now()

    # 센터를 '출발' 처리한 기사님의 배차 내역만 대기열로 가져옵니다.
    dispatches = Dispatch.query.filter(Dispatch.center_depart_time.isnot(None)).order_by(Dispatch.delivery_seq).all()
    
    data_list = []
    for d in dispatches:
        # 1. 드롭다운에서 선택한 특정 센터 필터링
        if center_name and d.center_name != center_name:
            continue
            
        # 2. 💡 [핵심 필터링] 토글이 ON일 때, 배송완료 건과 도착예정시간이 지난 매장 제외!
        if filter_past:
            if d.is_departed:
                continue
            if d.estimated_arrival and d.estimated_arrival < now:
                continue
        
        # 3. 빗금('/')이 포함된 다중 연락처 분할
        phones = []
        if d.store_phone:
            phones = [p.strip() for p in str(d.store_phone).split('/') if p.strip()]
        else:
            phones = ['']
            
        eta_str = d.estimated_arrival.strftime('%H:%M') if d.estimated_arrival else '계산중'
        template_str = d.template_name if d.template_name else '기본양식'
        
        # 연락처 개수만큼 행을 분할해서 저장
        for phone in phones:
            data_list.append({
                '센터명': d.center_name,
                '기사명': d.driver_name,
                '차량번호': d.vehicle_num,
                '매장명': d.store_name,
                '도착예상시간': eta_str,
                '수신번호(연락처)': phone,
                '적용양식(템플릿)': template_str
            })
    
    # 💡 만약 필터링 후 발송할 대상이 0건일 경우 에러 방지
    if not data_list:
        data_list.append({'안내': '해당 조건에 발송할 대상이 없습니다.'})

    # pandas를 이용해 엑셀 파일로 변환
    df = pd.DataFrame(data_list)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='SMS_발송대기열')
        
    output.seek(0)
    
    # 파일명 한글 깨짐 방지 인코딩
    today_str = datetime.now().strftime('%Y%m%d')
    filename = f"{center_filter}_{today_str}_알림톡발송양식.xlsx" if center_filter else f"{today_str}_알림톡발송양식.xlsx"
    
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/notice')
def notice_page():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    notices = Notice.query.order_by(Notice.display_seq.asc(), Notice.created_at.desc()).all()
    centers = Center.query.all() # 💡 [추가됨] 드롭다운용 센터 목록 불러오기
    
    comp_msg_setting = SystemSettings.query.filter_by(key='completion_msg').first()
    completion_message = comp_msg_setting.value if comp_msg_setting else "금일 배송도 고생 많으셨습니다!\n제때에서 발송된 카카오톡 배송승인 \n부탁드리겠습니다."
    return render_template('notice.html', notices=notices, centers=centers, completion_message=completion_message)

# 💡 [업데이트] 공지사항 이미지 업로드 (최대 5장) 처리 라우트
@app.route('/notice/add', methods=['POST'])
def add_notice():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    
    notice_id = request.form.get('notice_id') 
    title = request.form.get('title').strip()
    content = request.form.get('content').strip()
    
    # 💡 [신규] 센터명과 기사 조건을 '||' 기호로 묶어서 하나의 컬럼에 저장 (DB초기화 방지)
    target_center = request.form.get('target_center', '전체')
    raw_target_drivers = request.form.get('target_drivers', '').strip()
    combined_target = f"{target_center}||{raw_target_drivers}"
    
    display_seq = request.form.get('display_seq', type=int) or 1
    
    uploaded_images = []
    files = request.files.getlist('images')
    files = files[:5]
    for file in files:
        if file and file.filename != '' and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
            uploaded_images.append(unique_filename)
            
    images_str_val = "|".join(uploaded_images) if uploaded_images else None
    
    if notice_id:
        n = Notice.query.get(notice_id)
        if n:
            n.title = title
            n.content = content
            n.target_drivers = combined_target # 💡 묶은 데이터 저장
            n.display_seq = display_seq
            if images_str_val:
                n.images_str = images_str_val
    else:
        if title and content:
            db.session.add(Notice(title=title, content=content, images_str=images_str_val, target_drivers=combined_target, display_seq=display_seq, is_active=True))
            
    db.session.commit()
    return redirect(url_for('notice_page'))

@app.route('/notice/delete/<int:notice_id>', methods=['POST'])
def delete_notice(notice_id):
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    n = Notice.query.get(notice_id)
    if n:
        # 💡 보존을 위해 서버 저장 폴더 내 실물 이미지 파일도 자동 삭제
        for img in n.image_list:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], img))
            except Exception: pass
        db.session.delete(n)
        db.session.commit()
    return redirect(url_for('notice_page'))

@app.route('/notice/toggle/<int:notice_id>', methods=['POST'])
def toggle_notice(notice_id):
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    n = Notice.query.get(notice_id)
    if n:
        n.is_active = not n.is_active
        db.session.commit()
    return redirect(url_for('notice_page'))

# ==========================================
# 💡 [신규 추가] 매장별 문자 템플릿 양식 개별 수정
# ==========================================
@app.route('/admin/update_template/<int:dispatch_id>', methods=['POST'])
def update_template(dispatch_id):
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    dispatch = Dispatch.query.get(dispatch_id)
    if dispatch:
        template_name = request.form.get('template_name', '').strip()
        dispatch.template_name = template_name
        db.session.commit()
    return redirect(url_for('admin'))

# ==========================================
# 💡 퇴근(마지막 팝업) 텍스트 및 이미지 복사붙여넣기 등록
# ==========================================
@app.route('/upload_completion', methods=['POST'])
def upload_completion():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    
    # 1. 텍스트 문구 저장
    comp_text = request.form.get('completion_text', '').strip()
    if comp_text:
        setting = SystemSettings.query.filter_by(key='completion_msg').first()
        if setting:
            setting.value = comp_text
        else:
            db.session.add(SystemSettings(key='completion_msg', value=comp_text))
        db.session.commit()
        
    # 2. 이미지 저장 (복사/붙여넣기로 만들어진 가상의 파일 받기)
    file = request.files.get('completion_image')
    if file and file.filename != '':
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], 'completion.png')
        file.save(save_path)
        
    return redirect(url_for('notice_page'))

# ==========================================
# 💡 [수정] AI 최적 경로 정렬 (DB 센터 좌표 기준 탐색)
# ==========================================
@app.route('/optimize_route', methods=['POST'])
def optimize_route():
    driver_name = request.form.get('driver_name')
    if not driver_name:
        return redirect(url_for('driver'))

    dispatches = Dispatch.query.filter_by(driver_name=driver_name).order_by(Dispatch.delivery_seq).all()
    if not dispatches:
        return redirect(url_for('driver', driver_name=driver_name))

    completed = [d for d in dispatches if d.is_departed]
    uncompleted = [d for d in dispatches if not d.is_departed and d.store_x and d.store_y]
    uncompleted_no_coords = [d for d in dispatches if not d.is_departed and (not d.store_x or not d.store_y)]

    if not uncompleted:
        return redirect(url_for('driver', driver_name=driver_name))

    driver_center = dispatches[0].center_name
    
    # 💡 DB에 저장된 센터 좌표 가져오기
    target_center_obj = Center.query.filter_by(name=driver_center).first()
    if target_center_obj and target_center_obj.center_y and target_center_obj.center_x:
        current_lat, current_lng = float(target_center_obj.center_y), float(target_center_obj.center_x)
    else:
        current_lat, current_lng = 37.5665, 126.9780
    
    # 만약 이미 완료된 배송지가 있다면 마지막 완료지점을 기준으로 시작
    if completed:
        last_completed = completed[-1]
        if last_completed.store_y and last_completed.store_x:
            current_lat = float(last_completed.store_y)
            current_lng = float(last_completed.store_x)

    optimized_list = []
    current_node = (current_lat, current_lng)
    candidates = uncompleted.copy()
    
    while candidates:
        closest = min(candidates, key=lambda d: haversine(current_node[0], current_node[1], float(d.store_y), float(d.store_x)))
        optimized_list.append(closest)
        candidates.remove(closest)
        current_node = (float(closest.store_y), float(closest.store_x))

    start_seq = len(completed) + 1
    for idx, d in enumerate(optimized_list):
        d.delivery_seq = start_seq + idx
        
    for idx, d in enumerate(uncompleted_no_coords):
        d.delivery_seq = start_seq + len(optimized_list) + idx

    db.session.commit()
    return redirect(url_for('driver', driver_name=driver_name))

@app.route('/admin/bulk_update', methods=['POST'])
def bulk_update():
    if not session.get('is_admin'): return redirect(url_for('admin_login'))
    
    # 폼에 숨겨진 전체 dispatch ID 목록을 가져옵니다.
    dispatch_ids_str = request.form.get('dispatch_ids', '')
    if dispatch_ids_str:
        dispatch_ids = dispatch_ids_str.split(',')
        for did in dispatch_ids:
            if not did.strip(): continue
            d = Dispatch.query.get(did)
            if d:
                new_address = request.form.get(f'address_{did}', d.store_address).strip()
                
                # 💡 [핵심 수정] 주소가 기존과 달라졌거나, 기존에 좌표(x, y)가 없던 경우 카카오 API로 좌표 재탐색!
                if new_address != d.store_address or not d.store_x or not d.store_y:
                    d.store_address = new_address
                    c_x, c_y = get_kakao_coords(new_address)
                    if c_x and c_y:
                        d.store_x = c_x
                        d.store_y = c_y
                    else:
                        # 변환 실패 시 None 처리하여 관리자에게 다시 경고가 뜨도록 유지
                        d.store_x = None
                        d.store_y = None
                
                # 나머지 데이터 업데이트
                buffer_val = request.form.get(f'buffer_{did}')
                if buffer_val and buffer_val.isdigit():
                    d.buffer_time = int(buffer_val)
                    
                d.driver_phone = request.form.get(f'driver_phone_{did}', d.driver_phone or '').strip()
                d.store_phone = request.form.get(f'store_phone_{did}', d.store_phone or '').strip()
                d.driver_name = request.form.get(f'driver_name_{did}', d.driver_name).strip()
                d.vehicle_num = request.form.get(f'vehicle_num_{did}', d.vehicle_num).strip()
                d.template_name = request.form.get(f'template_{did}', d.template_name or '').strip()
                
        db.session.commit()
        
    return redirect(url_for('admin'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
